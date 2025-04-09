import xlrd
import openpyxl
import os
import tkinter as tk
from tkinter import filedialog

folder_path = r'E:\'
filename = "name.xls"
case = 0

def test_xls_file(xls_file, print_rep):
    hidden_sheets = False
    try:
        # Get the number of sheets
        num_sheets = xls_file.nsheets
        #get sheets and their names
        sheets = xls_file.sheets()
        sheet_names = xls_file.sheet_names()
        for sheet in sheets:
            visibility = sheet.visibility # 0 if is visible, 1 if is hidden and 2 if there are very hidden sheets
            if visibility != 0:
                hidden_sheets = True
                
        if print_rep == True:
            print(f"Number of sheets in the file: {num_sheets}")
            if hidden_sheets == True:
                print("There are hidden sheets!")
            else:
                print("No hidden sheets.")
            
    except Exception as e:
        print(f"An error occurred during the test of the xls file: {e}")

    return hidden_sheets

def what_hidden_sheets(xls_file):
    sheets = xls_file.sheets()
    print("---------------------------------------------------------------------")
    print("Sheets indexes, names and visibility status:")
    for index, sheet in enumerate(sheets):
        visibility = sheet.visibility
        status = "visible" if visibility == 0 else "hidden" if visibility == 1 else "very hidden"
        print(f"{index}, {sheet.name}: {status}")
    print("---------------------------------------------------------------------")

#if there are hidden sheets, it prints all the sheets
def indexes_visibile_sheets(xls_file):
    indexes = []
    sheets = xls_file.sheets()
    for index, sheet in enumerate(sheets):
        visibility = sheet.visibility
        status = "visible" if visibility == 0 else "hidden" if visibility == 1 else "very hidden"
        if visibility == 0:
            indexes.append(index)
    return indexes

def create_xlsx(xls_file):
    global workbook
    indexes = indexes_visibile_sheets(xls_file)
    sheets = xls_file.sheets()
    workbook = openpyxl.Workbook()
    # Remove the default sheet created by openpyxl
    workbook.remove(workbook['Sheet'])
    for index in indexes:
        sheet = xls_file.sheet_by_index(index)
        new_sheet = workbook.create_sheet(title = sheet.name) 
    return workbook

def save_xlsx(filename):
    file_path_xlsx = make_filename_xlsx(filename)
    workbook.save(file_path_xlsx)
    print("File saved in: ", file_path_xlsx)
    
def make_filename_xlsx(filename):
    filename_xlsx = filename[:-4] + '.xlsx'
    file_path_xlsx = os.path.join(folder_path, filename_xlsx)
    return file_path_xlsx

def choose_file():
    # Create a Tkinter root window
    root = tk.Tk()
    root.withdraw()  # Hide the root window

    # Open the file explorer window
    file_path = filedialog.askopenfilename()
    print(f"Selected file: {file_path}")

    return file_path


while case != 1 and case != 2:
    print("Do you want to manually choose the file or use the internal variable?(Type 1 or 2)")
    case = int(input())

if case == 1:
    file_path = choose_file()
else:
    file_path = os.path.join(folder_path, filename)


try:
    # Open the .xls file
    xls_file = xlrd.open_workbook(file_path)
    is_hidden_sheets = False
    is_hidden_sheets = test_xls_file(xls_file, False)
    
    if is_hidden_sheets == True:
        what_hidden_sheets(xls_file)

    workbook = create_xlsx(xls_file)


    # Copy data from .xls to .xlsx
    indexes = indexes_visibile_sheets(xls_file)
    for i,index in enumerate(indexes):
        sheet = xls_file.sheet_by_index(index)
        new_sheet = workbook.worksheets[i] 
    
        for row in range(sheet.nrows):
            for col in range(sheet.ncols):
                new_sheet.cell(row=row+1, column=col+1, value=sheet.cell_value(row, col))
                
    save_xlsx(filename)

    print("File converted successfully!")

except FileNotFoundError:
    print("The specified file was not found.")
except Exception as e:
    print(f"An error occurred: {e}")
